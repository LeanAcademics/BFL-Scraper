"""Phase 2 — populate the existing Mastersheet in
`label-consolidation/bfl-classification-consolidated.xlsx` with:

  * L0_label                                    (readable)
  * L1_summary                                  (readable, joined string)
  * L1_<cat>__<theme_slug>  × N                 (one-hot indicators)
  * L2_summary                                  (readable, joined string)
  * L2_<subcat-or-cat>__<theme_slug>  × N       (one-hot indicators)

All new columns are APPENDED to the right of the existing 23 — no existing
column is touched, reordered, or rewritten.

Safety:
  1. Timestamped backup is created before any write.
  2. Existing Mastersheet headers/data are read into memory before save and
     compared cell-by-cell against the saved file afterwards. If any
     discrepancy is found, the script aborts (you still have the backup).
  3. Indicator column sums are checked against the per-theme totals from
     Phase 1; mismatch → abort with details.
"""

import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "label-consolidation" / "bfl-classification-consolidated.xlsx"

EXPECTED_MASTERSHEET_HEADERS = [
    "BFL_id", "numeric_id", "L0_classification", "L1_act_count", "L1_categories",
    "L2_precondition_count", "L2_categories", "date", "age", "nationality",
    "location", "object_type", "discipline", "suit_raw", "base_seasons",
    "base_jumps", "ws_base_jumps", "skydives", "ws_skydives", "cause_of_death",
    "possible_factors", "weather", "accident",
]


def slug(s, max_len=40):
    s = str(s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s[:max_len] or "blank"


def rows_as_dicts(ws):
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    for row in rows:
        yield dict(zip(headers, row))


def snapshot_mastersheet(ws):
    """Read every cell of the Mastersheet's existing 23 columns into a list of
    tuples — to be compared after save."""
    snap = []
    for row in ws.iter_rows(min_row=1, max_col=len(EXPECTED_MASTERSHEET_HEADERS),
                            values_only=True):
        snap.append(tuple(row))
    return snap


def main():
    if not SRC.exists():
        sys.exit(f"Source not found: {SRC}")

    # ---- backup ----
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = SRC.parent / f"bfl-classification-consolidated.backup_{ts}.xlsx"
    shutil.copy2(SRC, backup)
    print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(SRC)
    if "Mastersheet" not in wb.sheetnames:
        sys.exit("'Mastersheet' not found")
    ms = wb["Mastersheet"]

    headers = [c.value for c in ms[1]]
    if headers != EXPECTED_MASTERSHEET_HEADERS:
        sys.exit(f"Mastersheet headers differ from expected.\nGot:      {headers}\nExpected: {EXPECTED_MASTERSHEET_HEADERS}")

    # Pre-save snapshot of the existing 23 columns
    pre_snapshot = snapshot_mastersheet(ms)
    print(f"Mastersheet rows (incl header): {len(pre_snapshot)}")

    # ---- build lookups ----
    l0_by_record = {}
    for r in rows_as_dicts(wb["Records"]):
        l0_by_record[r["BFL_id"]] = r.get("L0_label")

    l1_by_record = defaultdict(list)
    for r in rows_as_dicts(wb["L1_acts"]):
        l1_by_record[r["record_id"]].append({
            "category": r.get("L1_category"),
            "label": r.get("L1_label"),
        })

    l2_by_record = defaultdict(list)
    for r in rows_as_dicts(wb["L2_preconditions"]):
        l2_by_record[r["record_id"]].append({
            "category": r.get("L2_category"),
            "subcategory": r.get("L2_subcategory"),
            "label": r.get("L2_label"),
        })

    # ---- determine indicator columns (sorted for stable order) ----
    l1_keys = sorted({(a["category"], a["label"])
                      for acts in l1_by_record.values() for a in acts})
    l2_keys = sorted({((p["subcategory"] or p["category"]), p["label"])
                      for pcs in l2_by_record.values() for p in pcs})

    l1_columns = [(cat, lab, f"L1_{cat}__{slug(lab)}") for cat, lab in l1_keys]
    l2_columns = [(cat, lab, f"L2_{cat}__{slug(lab)}") for cat, lab in l2_keys]

    all_new_names = ["L0_label", "L1_summary"] + [c[2] for c in l1_columns] + \
                    ["L2_summary"] + [c[2] for c in l2_columns]
    if len(set(all_new_names)) != len(all_new_names):
        dupes = [n for n, c in Counter(all_new_names).items() if c > 1]
        sys.exit(f"Duplicate new column names: {dupes}")

    print(f"L1 indicator columns: {len(l1_columns)}")
    print(f"L2 indicator columns: {len(l2_columns)}")
    print(f"Total new columns:    {len(all_new_names)}")

    # ---- expected per-theme totals (sanity reference) ----
    expected_l1 = Counter((a["category"], a["label"])
                          for acts in l1_by_record.values() for a in acts)
    expected_l2 = Counter(((p["subcategory"] or p["category"]), p["label"])
                          for pcs in l2_by_record.values() for p in pcs)

    # ---- write headers (row 1) ----
    base = len(EXPECTED_MASTERSHEET_HEADERS)  # 23
    for i, name in enumerate(all_new_names):
        ms.cell(row=1, column=base + 1 + i, value=name)

    # column index ranges
    col_l0_label = base + 1
    col_l1_summary = col_l0_label + 1
    l1_ind_start = col_l1_summary + 1
    l1_ind_end = l1_ind_start + len(l1_columns) - 1
    col_l2_summary = l1_ind_end + 1
    l2_ind_start = col_l2_summary + 1
    l2_ind_end = l2_ind_start + len(l2_columns) - 1

    # ---- write per-record data ----
    n_records_seen = 0
    indicator_sums_l1 = Counter()
    indicator_sums_l2 = Counter()
    for r in range(2, ms.max_row + 1):
        bfl_id = ms.cell(row=r, column=1).value
        if bfl_id is None:
            continue
        n_records_seen += 1

        acts = l1_by_record.get(bfl_id, [])
        pcs = l2_by_record.get(bfl_id, [])

        l1_summary = " | ".join(f"[{a['category']}] {a['label']}" for a in acts)
        l2_summary = " | ".join(
            f"[{p['subcategory'] or p['category']}] {p['label']}" for p in pcs
        )

        ms.cell(row=r, column=col_l0_label, value=l0_by_record.get(bfl_id))
        ms.cell(row=r, column=col_l1_summary, value=l1_summary or None)

        record_l1_set = {(a["category"], a["label"]) for a in acts}
        for i, (cat, lab, _name) in enumerate(l1_columns):
            v = 1 if (cat, lab) in record_l1_set else 0
            ms.cell(row=r, column=l1_ind_start + i, value=v)
            if v:
                indicator_sums_l1[(cat, lab)] += 1

        ms.cell(row=r, column=col_l2_summary, value=l2_summary or None)
        record_l2_set = {((p["subcategory"] or p["category"]), p["label"]) for p in pcs}
        for i, (cat, lab, _name) in enumerate(l2_columns):
            v = 1 if (cat, lab) in record_l2_set else 0
            ms.cell(row=r, column=l2_ind_start + i, value=v)
            if v:
                indicator_sums_l2[(cat, lab)] += 1

    print(f"Records written: {n_records_seen}")

    # ---- pre-save sanity: indicator sums vs expected ----
    bad_l1 = [(k, indicator_sums_l1[k], expected_l1[k]) for k in expected_l1
              if indicator_sums_l1[k] != expected_l1[k]]
    bad_l2 = [(k, indicator_sums_l2[k], expected_l2[k]) for k in expected_l2
              if indicator_sums_l2[k] != expected_l2[k]]
    # NOTE: a record can have the same (cat, label) act twice → indicator stays 1
    # but expected counts each occurrence. We tolerate this case explicitly.
    def adjust_for_dupes(by_record, key_fn):
        adj = Counter()
        for items in by_record.values():
            keys = [key_fn(x) for x in items]
            for k in set(keys):  # dedup within record
                adj[k] += 1
        return adj
    expected_l1_dedup = adjust_for_dupes(l1_by_record, lambda a: (a["category"], a["label"]))
    expected_l2_dedup = adjust_for_dupes(l2_by_record, lambda p: ((p["subcategory"] or p["category"]), p["label"]))

    bad_l1 = [(k, indicator_sums_l1[k], expected_l1_dedup[k])
              for k in expected_l1_dedup if indicator_sums_l1[k] != expected_l1_dedup[k]]
    bad_l2 = [(k, indicator_sums_l2[k], expected_l2_dedup[k])
              for k in expected_l2_dedup if indicator_sums_l2[k] != expected_l2_dedup[k]]
    if bad_l1 or bad_l2:
        print("INDICATOR SUM MISMATCH (pre-save) — aborting before write")
        for k, got, exp in bad_l1: print(f"  L1 {k}: indicator sum={got} expected={exp}")
        for k, got, exp in bad_l2: print(f"  L2 {k}: indicator sum={got} expected={exp}")
        sys.exit(1)

    # ---- grouping (collapse indicator blocks by default) ----
    for c in range(l1_ind_start, l1_ind_end + 1):
        ms.column_dimensions[get_column_letter(c)].outline_level = 1
        ms.column_dimensions[get_column_letter(c)].hidden = True
    for c in range(l2_ind_start, l2_ind_end + 1):
        ms.column_dimensions[get_column_letter(c)].outline_level = 1
        ms.column_dimensions[get_column_letter(c)].hidden = True
    ms.sheet_properties.outlinePr.summaryRight = False

    ms.freeze_panes = "A2"
    ms.auto_filter.ref = f"A1:{get_column_letter(l2_ind_end)}{ms.max_row}"

    wb.save(SRC)
    print(f"Saved: {SRC}")

    # ---- POST-SAVE VERIFICATION: existing 23 columns must be byte-equal ----
    wb2 = openpyxl.load_workbook(SRC)
    ms2 = wb2["Mastersheet"]
    post_snapshot = snapshot_mastersheet(ms2)
    if pre_snapshot != post_snapshot:
        print("EXISTING-DATA MISMATCH AFTER SAVE — restoring backup")
        for i, (pre, post) in enumerate(zip(pre_snapshot, post_snapshot)):
            if pre != post:
                print(f"  row {i+1}: pre={pre} post={post}")
                if i > 5:
                    break
        shutil.copy2(backup, SRC)
        print(f"Restored from {backup}")
        sys.exit(1)

    # ---- POST-SAVE: verify per-theme indicator sums by reading the saved file ----
    print("\nPost-save verification: indicator-column sums by reading the saved file")
    headers2 = [c.value for c in ms2[1]]

    def col_idx(name):
        return headers2.index(name) + 1

    def column_sum(idx):
        return sum((ms2.cell(row=r, column=idx).value or 0) for r in range(2, ms2.max_row + 1))

    fail = False
    for cat, lab, name in l1_columns:
        got = column_sum(col_idx(name))
        exp = expected_l1_dedup[(cat, lab)]
        if got != exp:
            print(f"  FAIL L1 {name}: got={got} expected={exp}")
            fail = True
    for cat, lab, name in l2_columns:
        got = column_sum(col_idx(name))
        exp = expected_l2_dedup[(cat, lab)]
        if got != exp:
            print(f"  FAIL L2 {name}: got={got} expected={exp}")
            fail = True

    if fail:
        sys.exit("Indicator sum verification failed (post-save)")

    print("All checks passed.")
    print(f"\n  Existing 23 columns: byte-equal preserved")
    print(f"  New columns added:   {len(all_new_names)}")
    print(f"  Mastersheet width:   {len(EXPECTED_MASTERSHEET_HEADERS) + len(all_new_names)} columns")
    print(f"  Backup at:           {backup}")


if __name__ == "__main__":
    main()
